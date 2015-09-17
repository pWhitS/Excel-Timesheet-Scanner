from openpyxl import *   #import openpyxl into namespace
from employee import CEmployee
from config import CConfigParser
from sheetdata import CSheetData
import sys
import os

gWorkbook = ""
gWorksheet = ""
gEmployeeDict = {}


def getWorkbook(wbname):
	print("Loading Excel Workbook '" + wbname + "'...") 
	return load_workbook(filename = wbname)


def getWorksheet(sname):
	return gWorkbook[sname] 


def translateColumnNumber(colNum):
	alphaList = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
	             "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",]
	divn = colNum / 26
	modn = colNum % 26

	colStr = alphaList[divn-1]
	colStr += alphaList[modn-1]

	return colStr


def addNewEmployee(newEmployee):
	if newEmployee.name is None or newEmployee.name == "":
		return 1

	employeeKey = newEmployee.name.replace(" ", "").strip().upper()

	#if employee already exists, merge the hours
	if employeeKey in set(k.upper() for k in gEmployeeDict):
		employee = gEmployeeDict[employeeKey]
		employee.mergeHoursWith(newEmployee)
		return 2

	gEmployeeDict[employeeKey] = newEmployee
	return 0


def scanWorksheet(maxCell):
	iter_range = "A1:" + "B20" #maxCell
	#iter_range = "A11:AK11"
	for row in gWorksheet.iter_rows(iter_range): #'A6:AC6'
		newEmployee = CEmployee() #create new CEmployee instance

		for cell in row:		
			cellColor = cell.fill.fgColor.rgb
			cellValue = cell.value
			cellCol = cell.column
			#print gWorksheet.title, cellCol, cell.row, cellValue, cellColor
			newEmployee.setAttributes(cellColor, cellValue, cellCol)

		addNewEmployee(newEmployee)


def getConfiguration():
	confp = CConfigParser("ot_xlscan.conf")
	confp.readConfigs()
	return confp



if __name__ == "__main__":
	confp = getConfiguration()
	gWorkbook = getWorkbook(confp.excel_filename) 
	ofl = open(confp.output_filename, "w")

	#load special values into special dictionary
	sd = CSheetData() #get shared mutable objects
	for k in confp.specialCellsDict:
		tval = confp.specialCellsDict[k]
		sd.spValuesDict[k] = tval

	#load worksheets, find their maximum cell, then scan them
	for ws in confp.worksheetList:
		print ws
		gWorksheet = getWorksheet(ws.strip())
		maxRow = gWorksheet.max_row
		maxCol = gWorksheet.max_column
		maxCell = translateColumnNumber(maxCol) + str(maxRow)
		scanWorksheet(maxCell)


	for k in gEmployeeDict:
		ofl.write(gEmployeeDict[k].toString())
		ofl.write(gEmployeeDict[k].checkOvertime())
		ofl.write(gEmployeeDict[k].checkTotalHours())
		ofl.write("\n")

	'''
		if confp.output_all is True:
			outputDir = "output"
			if os.path.isdir(outputDir) is False:
				os.mkdir(outputDir)

			with open(outputDir + , "w")
	'''


	ofl.close()



